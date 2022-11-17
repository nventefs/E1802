from turtle import end_fill
from types import NoneType
import numpy as np
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment, Protection
#from openpyxl import load_workbook
import pandas as pd
import os
import datetime
from time import strftime

rows        = []
row         = []
filenames   = []
timestamps  = []
part_list   = []

# PARTS TO BE EVALUATED:

# 380LX152M250A052
# B0192-B
# TM8050H-8W
# MX1A-21NW
# MX1A-11NW
# B0203B
# STM32L011K4T6
# LM3488MM/NOPB

critical_parts = ['380LX152M250A052', 'B0192-BL', 'TM8050H-8W', 'MX1A-21NW', 'MX1A-11NW', 'B0203-BL', 'STM32L011K4T6', 'LM3488MM/NOPB']

def xlsx_name():
    directory = './Supply/4Front/Inventory Files/'
    global filenames
    global timestamps
    for filename in os.listdir(directory):
        f = os.path.join(directory, filename)
        d = os.path.getmtime(f)
        filenames.append(str(f[32:]))
        timestamps.append(str(d))

    for j in range(len(filenames)):
        index = j
        try:
            if timestamps[j + 1] > timestamps[j]:
                index = j + 1
        except:
            index = j
    return filenames[index]

def part_data(mfg_part_num, maxrow, sheet):
    for k in range(1,maxrow + 1):
        cell = sheet.cell(row = k, column = 9)
        if(str(cell.value) == mfg_part_num):
            part_no = str(mfg_part_num)
            cell = sheet.cell(row = k, column = 4)
            while cell.value == None:
                k = k - 1
                cell = sheet.cell(row = k, column = 4)
            qty_per = str(cell.value)
            cell = sheet.cell(row = k, column = 6)
            qty_on_hand = str(cell.value)
            return[part_no, qty_per, qty_on_hand]

def open_orders(mfg_part_num, maxrow, sheet):
    qty = []
    due = []
    for k in range(1,maxrow + 1):
        cell = sheet.cell(row = k, column = 9)
        if(str(cell.value) == mfg_part_num):
            qty.append(sheet.cell(row = k, column = 7).value)
            due.append(sheet.cell(row = k, column = 8).value)
    return [qty, due]

def excel_read(datafile): # TO-DO Build a structure/class for these and append values instead of individual variables
    partno_4Front = []
    datafile = './Supply/4Front/Inventory Files/' + datafile
    wb = openpyxl.load_workbook(datafile) 
    sheet = wb.active 
    maxrow = sheet.max_row
    for k in range(4,maxrow):
        cell = sheet.cell(row = k, column = 2)
        if (cell.value != None):
            if k != 50:
                if k != 74:
                    partno_4Front.append(cell.value)
    return [wb, sheet, maxrow, partno_4Front]

def get_parts(maxrow, sheet):
    part_list = []
    for k in range(3, maxrow + 1):
        cell = sheet.cell(row = k, column = 9)
        if str(cell.value) != 'None' or cell.value != None:
            part_list.append(cell.value)
    return(eliminate_duplicates(part_list))

def get_sheets(part_list):
    invalid_characters = ('\\', '/', '*', '?', ':', '[',']')        #  \ , / , * , ? , : , [ , ]
    sheet_name = []
    for part in part_list:
        for k in invalid_characters:
            if k in str(part):
                sheet_name.append(str(part).replace(k,' '))
                break
            else:
                sheet_name.append(str(part))
                break
    return(sheet_name)

def eliminate_duplicates(dataset):
    final_dataset = []
    for k in range(1,len(dataset)):
        if (dataset[k-1] != dataset[k]):
            final_dataset.append(dataset[k-1])
    return final_dataset

def populate_data(sheet, part, partno_4Front): #  part_list, maxrow_4Front, sht_4Front
    cell = sheet.cell(row = 1, column = 1)
    cell.value = "Part:"
    cell.font = Font(bold=True, color="c4262e")
    cell = sheet.cell(row = 1, column = 2)
    cell.value = str(part)

    cell = sheet.cell(row = 2, column = 1)
    cell.value = "Manufacturer:"
    cell.font = Font(bold=True, color="c4262e")
    cell = sheet.cell(row = 3, column = 1)
    cell.value = "4Front #:"
    cell.font = Font(bold=True, color="c4262e")
    cell = sheet.cell(row = 3, column = 2)
    cell.value = str(partno_4Front)
    
    cell = sheet.cell(row = 4, column = 1)
    cell.value = "MONTH"
    cell.font = Font(bold=True, color="ffffff")
    cell.fill = PatternFill("solid", fgColor="c4262e")
    cell = sheet.cell(row = 4, column = 2)
    cell.value = "INVENTORY"
    cell.font = Font(bold=True, color="ffffff")
    cell.fill = PatternFill("solid", fgColor="c4262e")
    cell = sheet.cell(row = 4, column = 3)
    cell.value = "ORDERS"
    cell.font = Font(bold=True, color="ffffff")
    cell.fill = PatternFill("solid", fgColor="c4262e")

    cell = sheet.cell(row = 5, column = 1)
    cell.value = "9/1/2022"
    cell.number_format = 'MM/DD/YYYY'
    cell.alignment = Alignment(horizontal="center", vertical="center")

    for k in range(6,45):
        cell = sheet.cell(row = k, column = 1)
        val = "A" + str(k-1)
        cell.value = "=DATE(YEAR(" + val + "),MONTH(" + val + ")+1,DAY(" + val + "))"
        cell.number_format = 'MM/DD/YYYY'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if k % 2 == 0:
            for j in range(1, 4):
                cell = sheet.cell(row = k, column = j)
                cell.fill = PatternFill("solid", fgColor="9a9b9c")
    return



file_4Front                                                 = xlsx_name()
[wbk_4Front, sht_4Front, maxrow_4Front, partno_4Front]      = excel_read(file_4Front)
part_list                                                   = get_parts(maxrow_4Front, sht_4Front)
[qty, due]                                                  = open_orders(part_list[16], maxrow_4Front, sht_4Front)
wbk                                                         = openpyxl.load_workbook('./Supply/4Front/Summary.xlsx')
sheet_names                                                 = get_sheets(part_list)

"""
At this point in the code, the following variables contain the following information:
- wbk_4Front: The workbook that 4Front sends us
- sht_4Front: The worksheets listed on the 4Front spreadsheet, typically only '4FS Planning Tool'
- max_row_4Front: The total number of rows on the 4Front spreadsheet
- part_list: the raw data part list from 4Front spreadsheet

- wbk: the exporting spreadsheet to submit the data to
- sheet_names: the adjusted sheet names where each sheet is a part
"""



print(sheet_names[41]) # test part of BZX84C18-7-F
wbksheet = wbk[sheet_names[41]]
populate_data(wbksheet, str(sheet_names[41]), partno_4Front[41])

wbk.save('./Supply/4Front/Summary.xlsx')
wbk.close()
wbk_4Front.close()